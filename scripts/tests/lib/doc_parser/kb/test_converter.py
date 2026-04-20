#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel sheet structure parser tests."""
import json
import pytest
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent.parent.parent.parent / "references" / "1.产品开发检查清单2025年.xlsx"


@pytest.fixture(scope="session")
def excel_workbook():
    """Session-scoped Excel workbook for preprocessor tests."""
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    yield wb
    wb.close()


def _find_sheet(wb, name_fragment):
    """Find a sheet by name fragment."""
    for name in wb.sheetnames:
        if name_fragment in name:
            return wb[name]
    return None


class TestSheetStructureParser:
    """Tests for parsing sheet structure from Excel."""

    def test_list_content_sheets(self):
        """Should return only content sheets (skip '分工' and '相关法规')."""
        pytest.importorskip("openpyxl")
        from lib.doc_parser.kb.converter.excel_to_md import _list_content_sheets

        sheets = _list_content_sheets(str(EXCEL_PATH))
        names = [s["name"] for s in sheets]
        assert "分工" not in names
        assert "相关法规" not in names
        assert len(sheets) == 11
        assert any("00" in n for n in names)

    def test_detect_regulation_boundaries_standard(self, excel_workbook):
        """Standard sheets (00, 02-05) have title in row 1, headers in row 2, regulation in row 3."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_sheet_structure

        sheet = _find_sheet(excel_workbook, "00")
        assert sheet is not None

        structure = parse_sheet_structure(sheet, "00. 对照样例")
        assert structure.header_row == 2
        assert structure.data_start_row == 4
        assert structure.regulation_name != ""

    def test_detect_regulation_boundaries_with_owner(self, excel_workbook):
        """Sheets with '产品开发责任人' row (01, 06-08) have headers in row 3, data in row 5."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_sheet_structure

        sheet = _find_sheet(excel_workbook, "01")
        assert sheet is not None

        structure = parse_sheet_structure(sheet, "01. 对照样例")
        assert structure.header_row == 3
        assert structure.data_start_row == 5

    def test_detect_sub_regulations_in_sheet_10(self, excel_workbook):
        """Sheet 10 has multiple regulation boundaries detected by non-numeric column A."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_sheet_structure

        sheet = _find_sheet(excel_workbook, "10")
        assert sheet is not None

        structure = parse_sheet_structure(sheet, "10. 对照样例")
        # Sheet 10 should have multiple sub-regulations
        assert len(structure.sub_regulations) >= 5

    def test_extract_metadata_columns(self, excel_workbook):
        """Should correctly identify metadata column indices (B-G)."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_sheet_structure

        sheet = _find_sheet(excel_workbook, "00")
        assert sheet is not None

        structure = parse_sheet_structure(sheet, "00. 对照样例")
        headers = structure.headers
        assert "项目" in headers.values() or any("项目" in str(v) for v in headers.values())


class TestClauseExtraction:
    """Tests for clause extraction and sub-regulation filtering."""

    def test_extract_clauses_returns_entries(self, excel_workbook):
        """Should extract non-empty clause entries with sequence numbers."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_sheet_structure, extract_clauses

        sheet = _find_sheet(excel_workbook, "00")
        assert sheet is not None

        structure = parse_sheet_structure(sheet, "00. 对照样例")
        clauses = extract_clauses(sheet, structure)
        assert len(clauses) > 0
        assert all(c.sequence > 0 for c in clauses)
        assert all(c.content for c in clauses)
        assert all(c.row > 0 for c in clauses)

    def test_extract_clauses_includes_metadata(self, excel_workbook):
        """Should extract metadata columns when present."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_sheet_structure, extract_clauses

        sheet = _find_sheet(excel_workbook, "00")
        assert sheet is not None

        structure = parse_sheet_structure(sheet, "00. 对照样例")
        clauses = extract_clauses(sheet, structure)
        clauses_with_meta = [c for c in clauses if c.metadata]
        assert len(clauses_with_meta) > 0

    def test_sheet_10_sub_regulation_filtering(self, excel_workbook):
        """Sheet 10 sub-regulations should produce separate clause groups."""
        from lib.doc_parser.kb.converter.excel_to_md import (
            parse_sheet_structure, extract_clauses, _filter_clauses_for_sub_reg,
        )

        sheet = _find_sheet(excel_workbook, "10")
        assert sheet is not None

        structure = parse_sheet_structure(sheet, "10. 对照样例")
        all_clauses = extract_clauses(sheet, structure)

        total_filtered = 0
        for sub_reg in structure.sub_regulations:
            filtered = _filter_clauses_for_sub_reg(all_clauses, sub_reg, structure)
            total_filtered += len(filtered)

        assert total_filtered == len(all_clauses)


class TestMarkdownGeneration:
    """Tests for Markdown output formatting."""

    def test_generate_frontmatter_yaml(self):
        """Should produce valid YAML frontmatter block."""
        from lib.doc_parser.kb.converter.excel_to_md import generate_frontmatter

        fm = generate_frontmatter(
            collection="00_保险法",
            regulation="保险法",
            tags=["保险法", "人身保险"],
        )
        assert fm.startswith("---\n")
        assert fm.endswith("---\n")
        assert "collection: 00_保险法" in fm
        assert "regulation: 保险法" in fm

    def test_clauses_to_markdown_structure(self):
        """Generated Markdown should have frontmatter, H1 title, and H2 clauses."""
        from lib.doc_parser.kb.converter.excel_to_md import clauses_to_markdown, ClauseEntry, generate_frontmatter

        fm = generate_frontmatter("01_负面清单检查", "负面清单", [])
        clauses = [
            ClauseEntry(sequence=1, content="第一条内容", metadata={"险种大类": "人身保险"}),
            ClauseEntry(sequence=2, content="第二条内容", metadata={}),
        ]
        md = clauses_to_markdown(clauses, fm, "负面清单")

        assert md.startswith("---\n")
        assert "# 负面清单" in md
        assert "## 第1项" in md
        assert "## 第2项" in md
        assert "> **元数据**: 险种大类=人身保险" in md
        assert "第一条内容" in md

    def test_clauses_to_markdown_no_metadata_no_blockquote(self):
        """Clauses without metadata should not produce blockquote lines."""
        from lib.doc_parser.kb.converter.excel_to_md import clauses_to_markdown, ClauseEntry, generate_frontmatter

        fm = generate_frontmatter("00_保险法", "保险法", [])
        clauses = [ClauseEntry(sequence=1, content="无标签条款")]
        md = clauses_to_markdown(clauses, fm, "保险法")

        assert "> **元数据**:" not in md

    def test_format_metadata_block(self):
        """Metadata block should use pipe separator."""
        from lib.doc_parser.kb.converter.excel_to_md import format_metadata_block

        result = format_metadata_block({"险种大类": "人身保险", "险种类型": "寿险"})
        assert "险种大类=人身保险 | 险种类型=寿险" in result

    def test_format_metadata_block_empty(self):
        """Empty metadata should return empty string."""
        from lib.doc_parser.kb.converter.excel_to_md import format_metadata_block

        assert format_metadata_block({}) == ""

    def test_safe_filename(self):
        """Should remove unsafe characters and replace spaces."""
        from lib.doc_parser.kb.converter.excel_to_md import _safe_filename

        assert _safe_filename('关于"分红险"的规定（2025年）') == "关于分红险的规定2025年"
        assert _safe_filename("  extra  spaces  ") == "extra_spaces"

    def test_simplify_regulation_name_strips_agency_and_doc_number(self):
        """Should remove 发文机关 and 文号 from regulation names."""
        from lib.doc_parser.kb.converter.excel_to_md import _simplify_regulation_name

        # Full name with agency and doc number
        result = _simplify_regulation_name(
            "中国银保监会办公厅关于进一步规范保险机构互联网人身保险业务有关事项的通知银保监办发2021108号"
        )
        assert result == "关于进一步规范保险机构互联网人身保险业务有关事项的通知"

    def test_simplify_regulation_name_preserves_topic(self):
        """Should not modify names that have no agency or doc number."""
        from lib.doc_parser.kb.converter.excel_to_md import _simplify_regulation_name

        assert _simplify_regulation_name("关于规范短期健康保险业务有关问题的通知") == \
            "关于规范短期健康保险业务有关问题的通知"
        assert _simplify_regulation_name("人身保险公司保险条款和保险费率管理办法") == \
            "人身保险公司保险条款和保险费率管理办法"
        assert _simplify_regulation_name("关于印发普通型人身保险精算规定") == \
            "关于印发普通型人身保险精算规定"

    def test_simplify_regulation_name_bracketed_doc_number(self):
        """Should remove bracketed doc numbers like （银保监办发〔2021〕7号）."""
        from lib.doc_parser.kb.converter.excel_to_md import _simplify_regulation_name

        result = _simplify_regulation_name(
            "中国银保监会办公厅关于规范短期健康保险业务有关问题的通知（银保监办发〔2021〕7号）"
        )
        assert "银保监办发" not in result
        assert "2021" not in result
        assert "关于规范短期健康保险业务有关问题的通知" in result

    def test_simplify_negative_list_name_with_version(self):
        """Should strip 负面清单 prefix and version, extract version to extra."""
        from lib.doc_parser.kb.converter.excel_to_md import _simplify_negative_list_name

        name, extra = _simplify_negative_list_name("\u201c负面清单\u201d2025版产品报送管理")
        assert name == "产品报送管理"
        assert extra == "2025版"

    def test_simplify_negative_list_name_with_parens(self):
        """Should handle （2025版）： format."""
        from lib.doc_parser.kb.converter.excel_to_md import _simplify_negative_list_name

        name, extra = _simplify_negative_list_name(
            "\u201c负面清单\u201d\uff082025版\uff09\uff1a产品条款表述"
        )
        assert name == "产品条款表述"
        assert extra == "2025版"

    def test_simplify_negative_list_name_preserves_topic(self):
        """Should not strip 负面清单 when it is part of the topic."""
        from lib.doc_parser.kb.converter.excel_to_md import _simplify_negative_list_name

        name, extra = _simplify_negative_list_name("负面清单未提及的银保监系统通报问题")
        assert name == "负面清单未提及的银保监系统通报问题"
        assert extra is None


class TestRegulationNameParser:
    """Tests for LLM-based regulation name parsing."""

    def test_parse_regulation_names_single(self):
        """Should parse single regulation with agency and doc number."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_regulation_names
        from unittest.mock import patch

        mock_response = [
            {
                "original": "中国银保监会办公厅关于规范短期健康保险业务有关问题的通知（银保监办发〔2021〕7号）",
                "short_name": "规范短期健康保险业务有关问题的通知",
                "agencies": ["中国银保监会办公厅"],
                "doc_numbers": ["银保监办发〔2021〕7号"],
            }
        ]

        with patch("lib.llm.factory.LLMClientFactory.create_name_parser_llm") as mock_factory:
            mock_instance = mock_factory.return_value
            mock_instance.generate.return_value = json.dumps(mock_response, ensure_ascii=False)
            mock_instance.close = lambda: None

            result = parse_regulation_names(
                ["中国银保监会办公厅关于规范短期健康保险业务有关问题的通知（银保监办发〔2021〕7号）"],
            )

        assert len(result) == 1
        info = list(result.values())[0]
        assert info["short_name"] == "规范短期健康保险业务有关问题的通知"
        assert info["agencies"] == ["中国银保监会办公厅"]
        assert info["doc_numbers"] == ["银保监办发〔2021〕7号"]

    def test_parse_regulation_names_multi_regulation(self):
        """Should handle multiple regulations merged with & separator."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_regulation_names
        from unittest.mock import patch

        mock_response = [
            {
                "original": "规范短期健康保险业务...（银保监办发〔2021〕7号）、续保表述...（电子报备系统通知公告2021-4-22）",
                "short_name": "规范短期健康保险业务有关问题的通知&短期健康险续保表述备案事项的通知",
                "agencies": ["中国银保监会办公厅", "电子报备系统"],
                "doc_numbers": ["银保监办发〔2021〕7号", "电子报备系统通知公告2021-4-22"],
            }
        ]

        with patch("lib.llm.factory.LLMClientFactory.create_name_parser_llm") as mock_factory:
            mock_instance = mock_factory.return_value
            mock_instance.generate.return_value = json.dumps(mock_response, ensure_ascii=False)
            mock_instance.close = lambda: None

            result = parse_regulation_names(
                ["规范短期健康保险业务...（银保监办发〔2021〕7号）、续保表述...（电子报备系统通知公告2021-4-22）"],
            )

        info = list(result.values())[0]
        assert "&" in info["short_name"]
        assert len(info["agencies"]) == 2
        assert len(info["doc_numbers"]) == 2

    def test_parse_regulation_names_no_doc_number(self):
        """Should handle regulations without doc numbers."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_regulation_names
        from unittest.mock import patch

        mock_response = [
            {
                "original": "其他检查",
                "short_name": "其他检查",
                "agencies": [],
                "doc_numbers": [],
            }
        ]

        with patch("lib.llm.factory.LLMClientFactory.create_name_parser_llm") as mock_factory:
            mock_instance = mock_factory.return_value
            mock_instance.generate.return_value = json.dumps(mock_response, ensure_ascii=False)
            mock_instance.close = lambda: None

            result = parse_regulation_names(["其他检查"])

        info = list(result.values())[0]
        assert info["agencies"] == []
        assert info["doc_numbers"] == []

    def test_parse_regulation_names_empty_input(self):
        """Should return empty dict for empty input."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_regulation_names

        result = parse_regulation_names([])
        assert result == {}

    def test_parse_regulation_names_llm_failure(self):
        """Should return empty dict when LLM returns invalid JSON."""
        from lib.doc_parser.kb.converter.excel_to_md import parse_regulation_names
        from unittest.mock import patch

        with patch("lib.llm.factory.LLMClientFactory.create_name_parser_llm") as mock_factory:
            mock_instance = mock_factory.return_value
            mock_instance.generate.return_value = "not json at all"
            mock_instance.close = lambda: None

            result = parse_regulation_names(["some regulation"])

        assert result == {}


class TestFrontmatterWithMetadata:
    """Tests for frontmatter with parsed regulation metadata."""

    def test_frontmatter_with_agencies_and_doc_numbers(self):
        """Should include 发文机关 and 文号 in frontmatter when parsed info provided."""
        from lib.doc_parser.kb.converter.excel_to_md import generate_frontmatter

        parsed_info = {
            "short_name": "规范短期健康保险业务有关问题的通知",
            "agencies": ["中国银保监会办公厅", "电子报备系统"],
            "doc_numbers": ["银保监办发〔2021〕7号", "电子报备系统通知公告2021-4-22"],
        }
        fm = generate_frontmatter(
            collection="06_短期健康保险",
            regulation="关于规范...",
            tags=["健康保险"],
            parsed_info=parsed_info,
        )
        assert "发文机关:" in fm
        assert "中国银保监会办公厅" in fm
        assert "电子报备系统" in fm
        assert "文号:" in fm
        assert "银保监办发" in fm
        assert "险种类型: 短期健康保险\n" in fm

    def test_frontmatter_without_parsed_info(self):
        """Should omit 发文机关 and 文号 when no parsed info."""
        from lib.doc_parser.kb.converter.excel_to_md import generate_frontmatter

        fm = generate_frontmatter(
            collection="01_负面清单检查",
            regulation="其他检查",
            tags=[],
        )
        assert "发文机关:" not in fm
        assert "文号:" not in fm
        assert "险种类型" not in fm
