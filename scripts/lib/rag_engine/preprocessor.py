#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel 产品开发检查清单 → Markdown 知识库转换脚本。

将 references/1.产品开发检查清单2025年.xlsx 转换为结构化 Markdown 知识库。
每个 sheet 按法规粒度拆分，提取元数据标签，处理内嵌表格图片。
"""
import argparse
import io
import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Optional

logger = logging.getLogger(__name__)

# 跳过的 sheet 名称
_SKIP_SHEETS = {"分工", "相关法规"}

# Sheet 序号 → 目录名映射（基于实际 Excel sheet 前缀编号）
_SHEET_DIR_MAP = {
    "00": "00_保险法",
    "01": "01_负面清单检查",
    "02": "02_条款费率管理办法",
    "03": "03_健康保险管理办法",
    "04": "04_普通型人身保险",
    "05": "05_分红型人身保险",
    "06": "06_短期健康保险",
    "07": "07_意外伤害保险",
    "08": "08_互联网保险产品",
    "09": "09_税优健康险",
    "10": "10_其他监管规定",
}

# 非险种目录（不提取险种类型到 frontmatter）
_NON_INSURANCE_TYPE_DIRS = {"00_保险法", "01_负面清单检查", "02_条款费率管理办法", "10_其他监管规定"}

_METADATA_COLUMNS = {
    3: "险种大类",
    4: "险种类型",
    5: "险种分型",
    6: "保险期限",
    7: "主附险",
}


@dataclass(frozen=True)
class SheetStructure:
    """解析后的 sheet 结构信息。"""
    sheet_name: str
    header_row: int
    data_start_row: int
    regulation_name: str
    headers: Dict[int, str] = field(default_factory=dict)
    sub_regulations: List[Dict] = field(default_factory=list)
    layout_type: str = "standard"


@dataclass(frozen=True)
class ClauseEntry:
    """单个检查条款。"""
    sequence: int
    content: str
    row: int = 0
    metadata: Dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class _ImageInfo:
    """Excel 内嵌图片信息。"""
    sheet_name: str
    row: int
    col: int
    image_data: bytes


def _get_sheet_code(sheet_name: str) -> str:
    """从 sheet 名称提取序号代码，如 '00. 对照"保险法"等法规检查' → '00'。"""
    match = re.match(r"(\d{2})", sheet_name)
    return match.group(1) if match else ""


def _get_dir_name(sheet_name: str) -> str:
    """获取 sheet 对应的目录名。"""
    code = _get_sheet_code(sheet_name)
    return _SHEET_DIR_MAP.get(code, f"{code}_unknown")


def _is_number(value) -> bool:
    """判断单元格值是否为纯数字（int 或数字字符串）。"""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str) and value.strip().isdigit():
        return True
    return False


def _list_content_sheets(excel_path: str) -> List[Dict]:
    """列出 Excel 中的内容 sheet（跳过'分工'和'相关法规'）。"""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    result = []
    for name in wb.sheetnames:
        if name in _SKIP_SHEETS:
            continue
        code = _get_sheet_code(name)
        if code:
            result.append({"name": name, "code": code, "dir": _get_dir_name(name)})
    wb.close()
    return result


def parse_sheet_structure(sheet, sheet_name: str) -> SheetStructure:
    """解析 sheet 的结构信息：header 行、数据起始行、法规名称、子法规边界。"""
    rows = list(sheet.iter_rows(min_row=1, max_row=6, values_only=True))

    layout_type = "standard"
    if len(rows) >= 2 and rows[1] and any(
        "产品开发责任人" in str(cell) for cell in rows[1] if cell
    ):
        layout_type = "with_owner"

    if layout_type == "standard":
        header_row = 2
        regulation_row = 3
        data_start_row = 4
    else:
        header_row = 3
        regulation_row = 4
        data_start_row = 5

    header_data = rows[header_row - 1] if len(rows) >= header_row else []
    headers = {}
    for idx, val in enumerate(header_data):
        if val:
            headers[idx] = str(val).strip()

    regulation_name = ""
    if len(rows) >= regulation_row and rows[regulation_row - 1]:
        regulation_name = str(rows[regulation_row - 1][0] or "").strip()
        if not regulation_name and len(rows) >= 1 and rows[0]:
            regulation_name = str(rows[0][0] or "").strip()

    regulation_name = re.sub(r'\n[（(][^）)]*[）)]', '', regulation_name)

    sub_regulations = []
    current_sub = {"name": regulation_name, "start_row": data_start_row}
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=data_start_row, values_only=False), start=data_start_row
    ):
        cell_a = row[0].value if row else None
        if cell_a is not None and not _is_number(cell_a):
            cell_a_str = str(cell_a).strip()
            if cell_a_str in ("序号",):
                continue
            if current_sub["name"] and current_sub["start_row"] != row_idx:
                sub_regulations.append(dict(current_sub))
            # 清理子法规名称：去掉换行后的括号注释（如 "\n（适用于互联网产品）"）
            clean_name = re.sub(r'\n[（(][^）)]*[）)]', '', cell_a_str)
            current_sub = {
                "name": clean_name,
                "start_row": row_idx + 1,            }
    if current_sub["name"]:
        sub_regulations.append(dict(current_sub))

    if len(sub_regulations) == 1 and sub_regulations[0]["name"] == regulation_name:
        sub_regulations = []

    return SheetStructure(
        sheet_name=sheet_name,
        header_row=header_row,
        data_start_row=data_start_row,
        regulation_name=regulation_name,
        headers=headers,
        sub_regulations=sub_regulations,
        layout_type=layout_type,
    )


def extract_clauses(sheet, structure: SheetStructure) -> List[ClauseEntry]:
    """从 sheet 中提取所有检查条款及其元数据。"""
    clauses = []

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=structure.data_start_row, values_only=True),
        start=structure.data_start_row,
    ):
        cell_a = row[0] if row else None

        if cell_a is None or (not _is_number(cell_a)):
            continue

        content = str(row[1] or "").strip() if len(row) > 1 else ""
        if not content:
            continue

        metadata = {}
        for col_idx, col_name in _METADATA_COLUMNS.items():
            if col_idx < len(row) and row[col_idx]:
                val = str(row[col_idx]).strip()
                if val and val != "全部":
                    metadata[col_name] = val

        clauses.append(ClauseEntry(
            sequence=int(cell_a) if isinstance(cell_a, (int, float)) else int(float(cell_a)),
            content=content,
            row=row_idx,
            metadata=metadata,
        ))

    return clauses


def format_metadata_block(metadata: Dict[str, str]) -> str:
    """将元数据字典格式化为 blockquote 格式。"""
    if not metadata:
        return ""
    parts = [f"{k}={v}" for k, v in metadata.items()]
    return f"\n> **元数据**: {' | '.join(parts)}\n"


def generate_frontmatter(
    collection: str,
    regulation: str,
    tags: List[str],
    parsed_info: Optional[dict] = None,
) -> str:
    """生成 YAML frontmatter。"""
    import yaml  # type: ignore[import-untyped]

    insurance_type = ""
    if collection and collection not in _NON_INSURANCE_TYPE_DIRS and "_" in collection:
        insurance_type = collection.split("_", 1)[1]

    data = {
        "collection": collection,
        "regulation": regulation,
        "tags": tags,
    }
    if insurance_type:
        data["险种类型"] = insurance_type
    if parsed_info:
        if parsed_info.get("agencies"):
            data["发文机关"] = parsed_info["agencies"]
        if parsed_info.get("doc_numbers"):
            data["文号"] = parsed_info["doc_numbers"]
        if parsed_info.get("extra_info"):
            data["备注"] = parsed_info["extra_info"]
    return "---\n" + yaml.dump(data, allow_unicode=True, default_flow_style=False) + "---\n"


def clauses_to_markdown(
    clauses: List[ClauseEntry],
    frontmatter: str,
    regulation_name: str,
) -> str:
    """将条款列表转换为 Markdown 文档内容。"""
    lines = [frontmatter, f"# {regulation_name}", ""]

    for clause in clauses:
        lines.append(f"## 第{clause.sequence}项")
        lines.append(format_metadata_block(clause.metadata))
        lines.append(clause.content)
        lines.append("")

    return "\n".join(lines)


def extract_images_from_sheet(sheet, sheet_name: str) -> List[_ImageInfo]:
    """从 sheet 中提取所有嵌入图片。"""
    images = []
    for img in sheet._images:
        anchor = img.anchor
        if hasattr(anchor, "_from"):
            row = anchor._from.row + 1
            col = anchor._from.col
        else:
            continue

        try:
            raw = img._data() if callable(getattr(img, "_data", None)) else None
            if raw:
                from PIL import Image as PILImage
                pil_img = PILImage.open(io.BytesIO(raw))
                img_buf = io.BytesIO()
                pil_img.save(img_buf, format="PNG")
                images.append(_ImageInfo(
                    sheet_name=sheet_name,
                    row=row,
                    col=col,
                    image_data=img_buf.getvalue(),
                ))
        except Exception as e:
            logger.warning(f"提取图片失败 sheet={sheet_name} row={row}: {e}")

    return images


def extract_images_from_excel(excel_path: str) -> List[_ImageInfo]:
    """从 Excel 中提取所有嵌入图片。"""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path)
    all_images = []
    for name in wb.sheetnames:
        if name in _SKIP_SHEETS:
            continue
        sheet = wb[name]
        images = extract_images_from_sheet(sheet, name)
        all_images.extend(images)
        logger.info(f"Sheet '{name}': 提取 {len(images)} 张图片")
    wb.close()
    return all_images


def ocr_image(image_data: bytes) -> str:
    """调用 OCR 识别图片中的表格。"""
    import base64
    from lib.llm import LLMClientFactory

    b64 = f"data:image/png;base64,{base64.b64encode(image_data).decode('utf-8')}"
    llm = LLMClientFactory.create_ocr_llm()
    try:
        result = llm.ocr_table(b64)
        return result
    finally:
        llm.close()



def _simplify_regulation_name(name: str) -> str:
    """从法规名称中移除发文机关和文号，保留主题名称。

    作为 LLM short_name 的后处理兜底：当 LLM 未能去除发文机关和文号时，
    用正则匹配常见模式进行清理。保留"关于"前缀和"的通知"等后缀。
    """
    clean = name

    clean = re.sub(r'[（(][^）)]*[发函公告号令][^）)]*[）)]', '', clean)
    clean = re.sub(
        r'(?:银保监|保监|金寿险|国金|金融|保险)?(?:办|会)?(?:发|函)\s*[〔\[]?\d+[\]〕]?\s*号',
        '', clean,
    )
    clean = re.sub(r'\d{4}年第\d+号', '', clean)
    clean = re.sub(r'通知公告\d{4}[-/]\d{1,2}[-/]\d{1,2}', '', clean)

    for pat in [
        r'中国(?:银行)?保险监督管理委员会(?:办公厅|人身险监管部)?',
        r'中国(?:银行)?保监会(?:办公厅)?',
        r'国家金融监督管理总局(?:人身险司)?',
        r'中国银保监会(?:办公厅)?',
        r'中国保监会(?:办公厅)?',
        r'原中国(?:银行)?保监会(?:办公厅)?',
        r'银保监会(?:办公厅)?',
        r'保监会(?:办公厅)?',
    ]:
        clean = re.sub(pat, '', clean)

    clean = re.sub(r'\s+', '', clean).strip('_-：: ')
    return clean if clean else name


def _simplify_negative_list_name(name: str) -> tuple:
    """简化负面清单法规名称：去掉"负面清单"前缀和引号，提取版本号。

    Returns:
        (simplified_name, extra_info_or_None)
        如 ('"负面清单"2025版产品报送管理', '2025版') → ('产品报送管理', '2025版')
    """
    extra = None
    ver_match = re.search(r'(\d{4})版', name)
    if ver_match:
        extra = f"{ver_match.group(1)}版"
    clean = name.replace('\u201c', '').replace('\u201d', '').replace('"', '')
    clean = re.sub(r'^负面清单[（(\d]', lambda m: m.group()[4:], clean)
    clean = re.sub(r'^负面清单\s*[:：]', '', clean)
    clean = re.sub(r'^[（(]?\d{4}版[）)]?\s*[:：]?\s*', '', clean)
    clean = clean.lstrip('：: ').strip()
    return (clean, extra)


def _extract_json_array(text: str) -> Optional[str]:
    """从 LLM 返回文本中提取 JSON 数组，处理 thinking 文本和 code block。"""
    start = text.find('[')
    if start == -1:
        return None
    depth = 0
    in_string = False
    escape_next = False
    for i in range(start, len(text)):
        ch = text[i]
        if escape_next:
            escape_next = False
            continue
        if ch == '\\' and in_string:
            escape_next = True
            continue
        if ch == '"' and not escape_next:
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == '[':
            depth += 1
        elif ch == ']':
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


def parse_regulation_names(
    regulations: List[str],
) -> Dict[str, dict]:
    """调用 LLM 批量解析法规名称，返回结构化结果。

    Args:
        regulations: regulation 名称列表

    Returns:
        dict keyed by original regulation name, each value:
        {
            "short_name": "主题1&主题2",
            "agencies": ["机关1", "机关2"],
            "doc_numbers": ["文号1", "文号2"]
        }
    """
    if not regulations:
        return {}

    numbered_list = "\n".join(
        f"{i + 1}. {reg}" for i, reg in enumerate(regulations)
    )

    prompt = f"""请解析以下保险法规名称列表，为每条提取结构化信息。

规则：
1. short_name: 法规主题名称（去掉发文机关和文号），多个法规用 & 拼接
2. agencies: 发文机关列表（每个法规对应一个，如"电子报备系统"也是发文机关）
3. doc_numbers: 发文字号列表（如 保监发〔2015〕93号、银保监规〔2022〕24号、2022年第8号、电子报备系统通知公告2023-8-22）
4. extra_info: 附加信息列表，如版本号"2025版"、适用范围等非主题信息

注意：
- 如果法规没有文号（如"其他检查"、"签字栏"），doc_numbers 和 agencies 都用空列表
- 《》书名号内的内容属于法规名称的一部分，不要去掉
- 严格返回 JSON 数组，每个元素包含 original, short_name, agencies, doc_numbers, extra_info 五个字段
- 不要返回任何多余文本，只返回 JSON

法规列表：
{numbered_list}"""

    from lib.llm import LLMClientFactory
    llm = LLMClientFactory.create_name_parser_llm()
    try:
        result = llm.generate(prompt, temperature=0.0, max_tokens=8192)
        json_str = _extract_json_array(result)
        if not json_str:
            logger.error("LLM 返回内容中未找到有效 JSON 数组")
            logger.debug(f"LLM response (first 500 chars): {result[:500]}")
            return {}
        items = json.loads(json_str)
        parsed = {}
        for item in items:
            original = item.get("original", "")
            parsed[original] = {
                "short_name": item.get("short_name", original),
                "agencies": item.get("agencies", []),
                "doc_numbers": item.get("doc_numbers", []),
                "extra_info": item.get("extra_info", []),
            }
        return parsed
    except (json.JSONDecodeError, KeyError, TypeError) as e:
        logger.error(f"解析 LLM 返回结果失败: {e}")
        return {}
    finally:
        llm.close()


def convert_excel_to_markdown(
    excel_path: str,
    output_dir: str,
    skip_ocr: bool = False,
) -> Path:
    """主转换函数：Excel → Markdown 知识库。"""
    import openpyxl

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    refs_dir = output_path

    images = extract_images_from_excel(excel_path) if not skip_ocr else []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    all_regulations = []
    for sheet_name in wb.sheetnames:
        if sheet_name in _SKIP_SHEETS:
            continue
        code = _get_sheet_code(sheet_name)
        if not code:
            continue
        dir_name = _get_dir_name(sheet_name)
        sheet = wb[sheet_name]
        structure = parse_sheet_structure(sheet, sheet_name)
        if structure.sub_regulations:
            for sub_reg in structure.sub_regulations:
                all_regulations.append((sub_reg["name"], sheet_name, dir_name, structure))
        else:
            all_regulations.append((structure.regulation_name, sheet_name, dir_name, structure))

    # 批量解析法规名称
    regulation_names = [r[0] for r in all_regulations if r[0]]
    parsed_map = {}
    if regulation_names:
        parsed_map = parse_regulation_names(regulation_names)
        logger.info(f"LLM 解析了 {len(parsed_map)}/{len(regulation_names)} 条法规名称")

    for d in refs_dir.iterdir():
        if d.is_dir() and d.name in _SHEET_DIR_MAP.values():
            for f in d.glob("*.md"):
                f.unlink()

    wb.close()
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    for reg_name, sheet_name, dir_name, structure in all_regulations:
        sheet_dir = refs_dir / dir_name
        sheet_dir.mkdir(parents=True, exist_ok=True)

        sheet = wb[sheet_name]
        tags = [reg_name] if reg_name else []
        parsed_info = parsed_map.get(reg_name)

        # 文件名：优先使用 LLM 解析的 short_name
        if parsed_info and parsed_info.get("short_name"):
            safe_name = parsed_info["short_name"]
        else:
            safe_name = reg_name

        # 负面清单检查目录：去掉"负面清单"前缀和引号，提取版本号到 extra_info
        if dir_name == "01_负面清单检查":
            safe_name, extra = _simplify_negative_list_name(safe_name)
            if extra and parsed_info:
                existing = parsed_info.get("extra_info", [])
                parsed_info["extra_info"] = list(set(existing + [extra]))

        safe_name = _simplify_regulation_name(safe_name)
        safe_name = _safe_filename(safe_name)

        clauses = extract_clauses(sheet, structure)
        if structure.sub_regulations:
            sub_reg = next(s for s in structure.sub_regulations if s["name"] == reg_name)
            clauses = _filter_clauses_for_sub_reg(clauses, sub_reg, structure)

        if not clauses:
            logger.debug(f"跳过无条款的法规: {reg_name}")
            continue

        fm = generate_frontmatter(dir_name, reg_name, tags, parsed_info)
        md_content = clauses_to_markdown(clauses, fm, reg_name)
        md_path = sheet_dir / f"{safe_name}.md"
        md_path.write_text(md_content, encoding="utf-8")
        logger.info(f"生成: {md_path} ({len(clauses)} 条)")

    wb.close()

    if images and not skip_ocr:
        _process_and_embed_images(images, refs_dir)

    md_count = len(list(refs_dir.rglob("*.md")))
    logger.info(f"生成完成: {md_count} 个文档")

    return output_path


def _filter_clauses_for_sub_reg(
    clauses: List[ClauseEntry], sub_reg: Dict, structure: SheetStructure
) -> List[ClauseEntry]:
    """过滤出属于指定子法规的条款（基于行号范围）。"""
    sub_idx = structure.sub_regulations.index(sub_reg)
    start = sub_reg["start_row"]
    end = (
        structure.sub_regulations[sub_idx + 1]["start_row"]
        if sub_idx + 1 < len(structure.sub_regulations)
        else float("inf")
    )
    return [c for c in clauses if start <= c.row < end]


def _safe_filename(name: str) -> str:
    """将法规名称转换为安全的文件名。"""
    name = re.sub(r'[<>:"/\\|?*（）【】《》、，。！？；：""\'\'…—〔〕「」]', '', name)
    name = re.sub(r'\s*&\s*', '&', name)
    name = re.sub(r'[\s]+', '_', name)
    max_bytes = 240
    encoded = name.encode("utf-8")
    if len(encoded) > max_bytes:
        encoded = encoded[:max_bytes]
        name = encoded.decode("utf-8", errors="ignore")
    return name.strip("_")


def _process_and_embed_images(
    images: List[_ImageInfo],
    refs_dir: Path,
) -> None:
    """OCR 处理图片并嵌入到对应的 Markdown 文件。"""
    for img_info in images:
        dir_name = _get_dir_name(img_info.sheet_name)
        sheet_dir = refs_dir / dir_name

        if not sheet_dir.exists():
            logger.warning(f"目录不存在: {sheet_dir}")
            continue

        try:
            md_table = ocr_image(img_info.image_data)
            if not md_table:
                logger.warning(f"OCR 返回空结果: sheet={img_info.sheet_name} row={img_info.row}")
                continue

            _embed_table_near_row(sheet_dir, img_info.row, md_table)
            logger.info(f"OCR 表格已嵌入: sheet={img_info.sheet_name} row={img_info.row}")

        except Exception as e:
            logger.error(f"OCR 处理失败: sheet={img_info.sheet_name} row={img_info.row}: {e}")


def _embed_table_near_row(sheet_dir: Path, row: int, md_table: str) -> None:
    """将 Markdown 表格嵌入到对应目录的文件中（追加到文件末尾）。"""
    md_files = list(sheet_dir.glob("*.md"))
    if not md_files:
        return

    target = md_files[0]
    table_section = f"\n## 费率表\n\n{md_table}\n"
    with target.open("a", encoding="utf-8") as f:
        f.write(table_section)


def main():
    parser = argparse.ArgumentParser(description="Excel 检查清单 → Markdown 知识库")
    parser.add_argument("--input", required=True, help="Excel 文件路径")
    parser.add_argument("--output", default=None, help="输出目录路径（默认为项目根目录 references/）")
    parser.add_argument("--skip-ocr", action="store_true", help="跳过 OCR 图片处理")
    args = parser.parse_args()

    output = args.output
    if not output:
        repo_root = Path(__file__).parent.parent.parent.parent
        output = str(repo_root / "references")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    convert_excel_to_markdown(
        excel_path=args.input,
        output_dir=output,
        skip_ocr=args.skip_ocr,
    )


if __name__ == "__main__":
    main()
